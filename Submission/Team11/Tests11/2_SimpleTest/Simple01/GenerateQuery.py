#!/usr/bin/python

import sys
from openpyxl import load_workbook

EXCEL_WORKBOOK_PATH = sys.argv[1]
OUTPUT_FILE_PATH = sys.argv[2]
SHEET = sys.argv[3]

INDEX_COLUMN = 'A'
COMMENT_COLUMN = 'E'
DECLARATION_COLUMN = 'B'
SELECT_COLUMN = 'C'
EXPECTED_ANSWER_COLUMN = 'D'
TIME_LIMIT = 5000

def main():
	workbook = load_workbook(filename = EXCEL_WORKBOOK_PATH)
	querySheet = workbook[SHEET]
		
	print('Generating Queries ... ')
	print('Workbook: ', EXCEL_WORKBOOK_PATH)
	print('Output  : ', OUTPUT_FILE_PATH)
	
	text = format(querySheet)
	
	createNewTextFile(text, OUTPUT_FILE_PATH)
	
	print('Done! :)')
	
def format(sheet):
	text = ''
	for rowId in range(2, sheet.max_row+1):
		text += getIndex(sheet, rowId) + ' - '
		text += getComment(sheet, rowId) + '\n'
		text += getDeclaration(sheet, rowId) + '\n'
		text += getSelect(sheet, rowId) + '\n'
		text += getExpectedAnswer(sheet, rowId) + '\n'
		text +=  str(TIME_LIMIT) + '\n'
	return text

def getIndex(sheet, rowId):
	blockId = INDEX_COLUMN + str(rowId)
	return str(sheet[blockId].value)
	
def getComment(sheet, rowId):
	blockId = COMMENT_COLUMN + str(rowId)
	return str(sheet[blockId].value)
	
def getDeclaration(sheet, rowId):
	blockId = DECLARATION_COLUMN + str(rowId)
	declaration = str(sheet[blockId].value)
	if (declaration == "NONE" or declaration == "None" or declaration == "none"):
		return ""
	return declaration
	
def getSelect(sheet, rowId):
	blockId = SELECT_COLUMN + str(rowId)
	select = str(sheet[blockId].value)
	if (select == "NONE" or select == "None" or select == "none"):
		return ""
	return select

def getExpectedAnswer(sheet, rowId):
	blockId = EXPECTED_ANSWER_COLUMN + str(rowId)
	expectedAns = str(sheet[blockId].value)
	if (expectedAns == "TRUE" or expectedAns == "True"):
		return "true"
	elif (expectedAns == "FALSE" or expectedAns == "False"):
		return "false"
	return expectedAns
	
def writeIntoTextFile(text, queryFilePath):
	file = open(queryFilePath, 'w')
	file.write(text)
	file.close()
	
def createNewTextFile(fileContent, outputFilePath):
	writeIntoTextFile(fileContent, outputFilePath)
	
if __name__ == '__main__':
	main()
	