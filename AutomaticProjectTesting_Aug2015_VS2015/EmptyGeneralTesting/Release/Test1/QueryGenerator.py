#!/usr/bin/python

from openpyxl import load_workbook

EXCEL_WORKBOOK_PATH = 'Test1/QuerySet1.xlsx'
OUTPUT_FILE_PATH = 'Test1/testQuery1.txt'

INDEX_COLUMN = 'A'
COMMENT_COLUMN = 'E'
DECLARATION_COLUMN = 'B'
SELECT_COLUMN = 'C'
EXPECTED_ANSWER_COLUMN = 'D'
TIME_LIMIT = 5000

def main():
	workbook = load_workbook(filename = EXCEL_WORKBOOK_PATH)
	querySheet1 = workbook['Complex']
	querySheet2 = workbook['Follows']
	querySheet3 = workbook['FollowsStar']
	querySheet4 = workbook['Parent']
	querySheet5 = workbook['ParentStar']
	
	text = ''
	text += format(querySheet1)
	text += format(querySheet2)
	text += format(querySheet3)
	text += format(querySheet4)
	text += format(querySheet5)
	
	createNewTextFile(text, OUTPUT_FILE_PATH)
	
	print('Done! :)')
	
def format(sheet):
	text = ''
	for rowId in range(2, sheet.max_row+1):
		text += getIndex(sheet, rowId) + ' - '
		text += getComment(sheet, rowId) + '\n'
		text += getDeclaration(sheet, rowId) + '\n'
		text += getSelect(sheet, rowId) + '\n'
		text += getExpectedAnswer(sheet, rowId) + '\n'
		text +=  str(TIME_LIMIT) + '\n'
	return text

def getIndex(sheet, rowId):
	blockId = INDEX_COLUMN + str(rowId)
	return str(sheet[blockId].value)
	
def getComment(sheet, rowId):
	blockId = COMMENT_COLUMN + str(rowId)
	return str(sheet[blockId].value)
	
def getDeclaration(sheet, rowId):
	blockId = DECLARATION_COLUMN + str(rowId)
	return str(sheet[blockId].value)
	
def getSelect(sheet, rowId):
	blockId = SELECT_COLUMN + str(rowId)
	return str(sheet[blockId].value)

def getExpectedAnswer(sheet, rowId):
	blockId = EXPECTED_ANSWER_COLUMN + str(rowId)
	return str(sheet[blockId].value)
	
def writeIntoTextFile(text, queryFilePath):
	file = open(queryFilePath, 'w')
	file.write(text)
	file.close()
	
def createNewTextFile(fileContent, outputFilePath):
	writeIntoTextFile(fileContent, outputFilePath)
	
if __name__ == '__main__':
	main()
	