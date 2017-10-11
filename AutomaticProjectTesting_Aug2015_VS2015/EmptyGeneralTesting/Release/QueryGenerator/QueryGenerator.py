#!/usr/bin/python

from openpyxl import load_workbook

EXCEL_WORKBOOK_PATH = 'Queries.xlsx'

INDEX_COLUMN = 'A'
COMMENT_COLUMN = 'E'
DECLARATION_COLUMN = 'B'
SELECT_COLUMN = 'C'
EXPECTED_ANSWER_COLUMN = 'D'
TIME_LIMIT = 5000

def main():
	workbook = load_workbook(filename = EXCEL_WORKBOOK_PATH)
	querySheet1 = workbook['Queries1']
	querySheet2 = workbook['Queries2']
	querySheet3 = workbook['Queries3']
	
	createNewTextFile(querySheet1, 'testQuery1.txt')
	createNewTextFile(querySheet2, 'testQuery2.txt')
	createNewTextFile(querySheet3, 'testQuery3.txt')
	
	print('Done! :)')
	
def format(sheet):
	text = ''
	for rowId in range(2, sheet.max_row+1):
		text += getIndex(sheet, rowId) + ' - '
		text += getComment(sheet, rowId) + '\n'
		text += getDeclaration(sheet, rowId) + '\n'
		text += getSelect(sheet, rowId) + '\n'
		text += getExpectedAnswer(sheet, rowId) + '\n'
		text +=  str(TIME_LIMIT) + '\n'
	return text

def getIndex(sheet, rowId):
	blockId = INDEX_COLUMN + str(rowId)
	return str(sheet[blockId].value)
	
def getComment(sheet, rowId):
	blockId = COMMENT_COLUMN + str(rowId)
	return str(sheet[blockId].value)
	
def getDeclaration(sheet, rowId):
	blockId = DECLARATION_COLUMN + str(rowId)
	return str(sheet[blockId].value)
	
def getSelect(sheet, rowId):
	blockId = SELECT_COLUMN + str(rowId)
	return str(sheet[blockId].value)

def getExpectedAnswer(sheet, rowId):
	blockId = EXPECTED_ANSWER_COLUMN + str(rowId)
	return str(sheet[blockId].value)
	
def writeIntoTextFile(text, queryFilePath):
	file = open(queryFilePath, 'w')
	file.write(text)
	file.close()
	
def createNewTextFile(sheet, outputFilePath):
	fileContent = format(sheet)
	writeIntoTextFile(fileContent, outputFilePath)
	
if __name__ == '__main__':
	main()
	